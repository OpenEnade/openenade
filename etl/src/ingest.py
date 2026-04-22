"""Ingest ENADE XLSX files from INEP."""

from pathlib import Path

import openpyxl
import polars as pl

# Known sheet names for the main data across years
_DATA_SHEET_NAMES = {"PLANILHA_ENADE", "Planilha1"}

# Some years use non-standard sheet names. Match by substring.
_DATA_SHEET_PATTERNS = ["conceito enade", "planilha_enade", "planilha1", "enade"]


def read_enade_xlsx(path: Path) -> pl.DataFrame:
    """Read an ENADE Conceito XLSX file and return a Polars DataFrame.

    Finds the main data sheet by name (varies across years),
    reads all rows, and returns them as a DataFrame.
    """
    wb = openpyxl.load_workbook(path, read_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name in _DATA_SHEET_NAMES:
            sheet_name = name
            break
    # Fallback: match by substring pattern
    if sheet_name is None:
        for name in wb.sheetnames:
            name_lower = name.lower().strip()
            for pattern in _DATA_SHEET_PATTERNS:
                if pattern in name_lower:
                    sheet_name = name
                    break
            if sheet_name:
                break

    if sheet_name is None:
        wb.close()
        raise ValueError(
            f"No data sheet found in {path.name}. "
            f"Expected one of: {_DATA_SHEET_NAMES}. "
            f"Found: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return pl.DataFrame()

    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
    data_rows = rows[1:]

    # Filter out footnote/metadata rows: valid data rows have a numeric year
    # in the first column (Ano). Footnotes have text strings instead.
    clean_rows = []
    for row in data_rows:
        first_val = row[0]
        if first_val is not None and isinstance(first_val, (int, float)):
            clean_rows.append(row)

    # Convert empty strings to None so Polars can infer types correctly
    clean_data = {}
    for i, h in enumerate(headers):
        values = []
        for row in clean_rows:
            val = row[i]
            if val == "" or val == "-":
                val = None
            values.append(val)
        clean_data[h] = values

    return pl.DataFrame(clean_data, strict=False)
