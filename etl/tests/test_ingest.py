from pathlib import Path

import polars as pl

from src.ingest import read_enade_xlsx


def test_read_enade_xlsx_returns_dataframe(tmp_path: Path):
    """read_enade_xlsx should return a DataFrame with rows from the main data sheet."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLANILHA_ENADE"
    ws.append(["Ano", "Área de Avaliação", "Conceito Enade (Faixa)"])
    ws.append([2023, "MEDICINA", "4"])
    ws.append([2023, "MEDICINA", "SC"])
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)

    df = read_enade_xlsx(xlsx_path)

    assert isinstance(df, pl.DataFrame)
    assert len(df) == 2
    assert "Ano" in df.columns
    assert "Área de Avaliação" in df.columns


def test_read_enade_xlsx_finds_planilha1(tmp_path: Path):
    """Should also find the data sheet when it is named 'Planilha1'."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    ws.append(["Ano", "Conceito Enade (Faixa)"])
    ws.append([2022, "3"])
    xlsx_path = tmp_path / "test2.xlsx"
    wb.save(xlsx_path)

    df = read_enade_xlsx(xlsx_path)

    assert len(df) == 1


def test_read_enade_xlsx_raises_on_no_data_sheet(tmp_path: Path):
    """Should raise ValueError when no known data sheet is found."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RandomSheet"
    ws.append(["foo", "bar"])
    xlsx_path = tmp_path / "bad.xlsx"
    wb.save(xlsx_path)

    import pytest

    with pytest.raises(ValueError, match="No data sheet found"):
        read_enade_xlsx(xlsx_path)


def test_read_enade_xlsx_finds_conceito_enade_sheet(tmp_path: Path):
    """Should find 'Conceito Enade2019' via substring pattern matching (2019 format)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conceito Enade2019"
    ws.append(["Ano", "Área de Avaliação"])
    ws.append([2019, "MEDICINA"])
    xlsx_path = tmp_path / "test_2019.xlsx"
    wb.save(xlsx_path)

    df = read_enade_xlsx(xlsx_path)

    assert len(df) == 1


def test_read_enade_xlsx_filters_footnote_rows(tmp_path: Path):
    """Rows where the first column is not a number (footnotes) should be filtered out."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLANILHA_ENADE"
    ws.append(["Ano", "Área de Avaliação", "Conceito Enade (Faixa)"])
    ws.append([2023, "MEDICINA", "4"])
    ws.append(["Fonte: INEP/MEC", None, None])  # footnote row
    ws.append([None, None, None])  # empty row
    xlsx_path = tmp_path / "with_footnotes.xlsx"
    wb.save(xlsx_path)

    df = read_enade_xlsx(xlsx_path)

    assert len(df) == 1
    assert df["Ano"][0] == 2023


def test_read_enade_xlsx_converts_dash_to_none(tmp_path: Path):
    """Cells with '-' should become None."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLANILHA_ENADE"
    ws.append(["Ano", "CEBAS"])
    ws.append([2023, "-"])
    xlsx_path = tmp_path / "with_dash.xlsx"
    wb.save(xlsx_path)

    df = read_enade_xlsx(xlsx_path)

    assert df["CEBAS"][0] is None


def test_read_enade_xlsx_returns_empty_for_header_only(tmp_path: Path):
    """XLSX with only a header row (no data) should return empty DataFrame."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLANILHA_ENADE"
    ws.append(["Ano", "Área de Avaliação"])
    xlsx_path = tmp_path / "header_only.xlsx"
    wb.save(xlsx_path)

    df = read_enade_xlsx(xlsx_path)

    assert len(df) == 0
